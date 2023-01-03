import requests
import time
from datetime import datetime
from openpyxl import *
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.ensemble import ExtraTreesClassifier

def QualCor(aux):
    if  aux == 0: return 'BRANCO'
    elif aux == 1: return 'VERMELHO'
    elif aux == 2: return 'PRETO'

def SalvarDadosUltimasRodadas():
    url = 'https://blaze.com/api/roulette_games/recent'
    try:
        response = requests.get(url)
        r = response.json()
    except:pass
        
    hora = float(((datetime.now()).strftime('%H')))
    minuto = float(((datetime.now()).strftime('%M')))
    try: 
        wb = load_workbook(filename='bd.xlsx')
        sheet = wb.active
        count = 3
        for rows in sheet.iter_rows(min_row=2):
            count += 1
        count -= 1
        sheet.cell(row=count, column=21).value = minuto
        sheet.cell(row=count, column=20).value = hora
        sheet.cell(row=count, column=19).value = r[0]['color']
        sheet.cell(row=count, column=18).value = r[1]['color']
        sheet.cell(row=count, column=17).value = r[2]['color']
        sheet.cell(row=count, column=16).value = r[3]['color']
        sheet.cell(row=count, column=15).value = r[4]['color']
        sheet.cell(row=count, column=14).value = r[5]['color']
        sheet.cell(row=count, column=13).value = r[6]['color']
        sheet.cell(row=count, column=12).value = r[7]['color']
        sheet.cell(row=count, column=11).value = r[8]['color']
        sheet.cell(row=count, column=10).value = r[9]['color']
        sheet.cell(row=count, column=9).value = r[10]['color']
        sheet.cell(row=count, column=8).value = r[11]['color']
        sheet.cell(row=count, column=7).value = r[12]['color']
        sheet.cell(row=count, column=6).value = r[13]['color']
        sheet.cell(row=count, column=5).value = r[14]['color']
        sheet.cell(row=count, column=4).value = r[15]['color']
        sheet.cell(row=count, column=3).value = r[16]['color']
        sheet.cell(row=count, column=2).value = r[17]['color']
        sheet.cell(row=count, column=1).value = r[18]['color']

        wb.save(filename='bd.xlsx')
        wb.close
        
    except: pass
    
def SalvarDadosRodadaAtual():
    url = 'https://blaze.com/api/roulette_games/recent'
    try:
        response = requests.get(url)
        r = response.json()
    except:pass
    idAnterior = r[0]['id']
    while idAnterior == r[0]['id']:
        try:
            response = requests.get(url)
            r = response.json()
        except: pass
    
    try:     
        wb = load_workbook(filename='bd.xlsx')
        sheet = wb.active
        count = 2
        for rows in sheet.iter_rows(min_row=2):
            count += 1
        count -=1
        sheet.cell(row=count, column=22).value = r[0]['color']
        wb.save(filename='bd.xlsx')
        wb.close
    except: pass
    # print('Dados Salvos')


url = 'https://blaze.com/api/roulette_games/recent'
parcialWin = 0
parcialLoss = 0
auxMaxLoss = 0
auxMaxWin = 0
maxLoss = 0
maxWin = 0

while True:
    try:
        response = requests.get(url)
        r = response.json()
    except:pass    
    arquivo = pd.read_excel('bd.xlsx')
    
    # df = pd.DataFrame(arquivo.loc[(arquivo['RODADA_1']== r[18]["color"]) 
    #                               & (arquivo['RODADA_2']== r[17]["color"])
    #                               & (arquivo['RODADA_3']== r[16]["color"])
    #                               & (arquivo['RODADA_4']== r[15]["color"]) 
    #                               & (arquivo['RODADA_5']== r[14]["color"]) 
    #                               & (arquivo['RODADA_6']== r[13]["color"]) 
    #                               & (arquivo['RODADA_7']== r[12]["color"]) 
    #                               & (arquivo['RODADA_8']== r[11]["color"]) 
    #                               & (arquivo['RODADA_9']== r[10]["color"]) 
    #                               & (arquivo['RODADA_10']== r[9]["color"]) 
    #                               & (arquivo['RODADA_11']== r[8]["color"]) 
    #                               & (arquivo['RODADA_12']== r[7]["color"]) 
    #                               & (arquivo['RODADA_13']== r[6]["color"]) 
    #                               & (arquivo['RODADA_14']== r[5]["color"])
    #                               & (arquivo['RODADA_15']== r[4]["color"]) 
    #                               & (arquivo['RODADA_16']== r[3]["color"]) 
    #                               & (arquivo['RODADA_17']== r[2]["color"]) 
    #                               & (arquivo['RODADA_18']== r[1]["color"])
    #                               & (arquivo['RODADA_19']== r[0]["color"]) ])
                                   
                                  
    # df = pd.DataFrame(arquivo.loc[(arquivo['RODADA_15'] == r[4]["color"]) 
    #                               & (arquivo['RODADA_16'] == r[3]["color"]) 
    #                               & (arquivo['RODADA_17'] == r[2]["color"]) 
    #                               & (arquivo['RODADA_18'] == r[1]["color"]) 
    #                               & (arquivo['RODADA_19'] == r[0]["color"]) ])
    
    df = pd.DataFrame(arquivo.loc[(arquivo['RODADA_10']== r[9]["color"]) 
                                    & (arquivo['RODADA_11'] == r[8]["color"])
                                    & (arquivo['RODADA_12'] == r[7]["color"]) 
                                    & (arquivo['RODADA_13'] == r[6]["color"]) 
                                    & (arquivo['RODADA_14'] == r[5]["color"])
                                    & (arquivo['RODADA_15'] == r[4]["color"]) 
                                    & (arquivo['RODADA_16'] == r[3]["color"]) 
                                    & (arquivo['RODADA_17'] == r[2]["color"]) 
                                    & (arquivo['RODADA_18'] == r[1]["color"])
                                    & (arquivo['RODADA_19'] == r[0]["color"]) ])
    
    
    # df = pd.DataFrame(arquivo.loc[(arquivo['RODADA_18'] == r[1]["color"]) 
    #                               & (arquivo['RODADA_19'] == r[0]["color"]) ])
    
    
    # df = pd.DataFrame(arquivo.loc[(arquivo['RODADA_19'] == r[0]["color"]) ])
    
    if len(df) > 0:
        print('\n\n')
        print(70 * '-')
        # print(df)
        # print(r[4]["color"],r[3]["color"],r[2]["color"],r[1]["color"],r[0]["color"])
        # print('* Padrão encontrado *')

        branco = round(100*(len(df[df['COR_PREVISTA'] == 0])/len(df)),2)
        vermelho = round(100*(len(df[df['COR_PREVISTA'] == 1])/len(df)),2)
        preto = round(100*(len(df[df['COR_PREVISTA'] == 2])/len(df)),2)
        
        # print(f'->{branco}% de BRANCO')
        print(f'->{vermelho}% de VERMELHO')
        print(f'->{preto}% de PRETO')
        
        # if branco > vermelho and branco > preto: corPrevista = 'BRANCO'
        if preto > branco and preto > vermelho: corPrevista = 'VERMELHO' #'PRETO'# # 
        elif vermelho > branco and vermelho > preto: corPrevista ='PRETO' #'PRETO' ## 
        else: corPrevista = ''
        
        if corPrevista != '':print(f'COR PREVISTA: {corPrevista}')
        else: print('SEM COR PREVISTA')
        SalvarDadosUltimasRodadas()
        SalvarDadosRodadaAtual()
        try:
            response = requests.get(url)
            r = response.json()
        except:pass
        if corPrevista != '':
            if QualCor(r[0]['color']) == corPrevista:
                auxMaxLoss = 0
                parcialWin += 1
                auxMaxWin += 1
                if maxWin < auxMaxWin: maxWin = auxMaxWin
                print('-> ✅ Green')
            else:
                auxMaxWin = 0
                parcialLoss += 1
                auxMaxLoss += 1
                if maxLoss < auxMaxLoss: maxLoss = auxMaxLoss
                print('-> ⛔️ Loss')
            winrate = round(100*(parcialWin / (parcialWin + parcialLoss)),2)
            msg = f'-> Parcial: ✅ = {parcialWin} | ⛔️ = {parcialLoss} | 🎯 = {winrate}% de acerto'
            print(msg)
            print(f'-> Máximo de VITORIA seguida: {maxWin}')
            print(f'-> Máximo de DERROTA seguida: {maxLoss}')

    else: 
        print('padrão NÃO encontrado, coletando dados.')
        SalvarDadosUltimasRodadas()
        SalvarDadosRodadaAtual()

    
    
    
    